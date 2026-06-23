import openpyxl

# 1. Parse Markdown file to split into June 13 and June 14 lists
md_path = "/Users/nicksng/code/puzzle-portal/ocr_tools/results/attendance_transcription.md"
with open(md_path, "r", encoding="utf-8") as f:
    content = f.read()

june13_rows = []
june14_rows = []
current_section = None

for line in content.split("\n"):
    line = line.strip()
    if "## 13-June-2026" in line:
        current_section = 13
        continue
    elif "## 14-June-2026" in line:
        current_section = 14
        continue
    
    if not line.startswith("|"):
        continue
    parts = [p.strip() for p in line.split("|")[1:-1]]
    if not parts or parts[0] == "ល.រ" or parts[0].startswith("---") or parts[0].startswith(":---"):
        continue
    
    no_val = parts[0]
    name = parts[1]
    gender = parts[2]
    role = parts[3]
    company = parts[4]
    phone = parts[5]
    
    # Translate Gender to Status (ស្រី or ប្រុស)
    if gender in ["F", "F ", "ស្រី"]:
        status = "ស្រី"
    elif gender in ["M", "M ", "ប្រុស"]:
        status = "ប្រុស"
    else:
        status = gender

    row_data = {
        "num": int(no_val) if no_val.isdigit() else no_val,
        "name": name,
        "status": status,
        "role": role if role not in ["None", ""] else None,
        "company": company if company not in ["None", ""] else None,
        "phone": phone if phone not in ["None", ""] else None
    }
    
    if current_section == 13:
        june13_rows.append(row_data)
    elif current_section == 14:
        june14_rows.append(row_data)

print(f"Parsed {len(june13_rows)} rows for June 13 and {len(june14_rows)} rows for June 14.")

# 2. Load Excel template
wb = openpyxl.load_workbook("/Users/nicksng/code/puzzle-portal/ocr_tools/results/sheet_export.xlsx")

# Get template sheet
template_sheet = wb["Actual Attendance MDK"]

# Create 13-June-2026 tab by copying the template
sheet_13 = wb.copy_worksheet(template_sheet)
sheet_13.title = "13-June-2026"

# Create 14-June-2026 tab by copying the template
sheet_14 = wb.copy_worksheet(template_sheet)
sheet_14.title = "14-June-2026"

# Remove the original template sheet so it's a clean new file
wb.remove(template_sheet)

def populate_sheet(sheet, rows, total_count):
    # Clear the original data rows in the template (from row 7 up to 330)
    for r in range(7, 331):
        for c in range(2, 9):
            sheet.cell(row=r, column=c, value=None)
            
    # Update formulas in Row 1, 3, 4
    # Row 1: Guests count formula
    sheet.cell(row=1, column=3, value=f'="Guests ("&SUM(D7:D{7 + total_count - 1})&")"')
    # Row 3: Male count formula and percentage
    sheet.cell(row=3, column=4, value=f'=COUNTIF($E$7:$E${7 + total_count - 1}, "ប្រុស")')
    sheet.cell(row=3, column=5, value=f'=D3/{total_count}')
    # Row 4: Female count formula and percentage
    sheet.cell(row=4, column=4, value=f'=COUNTIF($E$7:$E${7 + total_count - 1}, "ស្រី")')
    sheet.cell(row=4, column=5, value=f'=D4/{total_count}')
    
    # Write the new data rows starting at Row 7
    start_row = 7
    for i, row in enumerate(rows):
        current_row = start_row + i
        sheet.cell(row=current_row, column=2, value=row["num"])
        sheet.cell(row=current_row, column=3, value=row["name"])
        sheet.cell(row=current_row, column=4, value=None)  # Qty is blank
        sheet.cell(row=current_row, column=5, value=row["status"])
        sheet.cell(row=current_row, column=6, value=row["role"])
        sheet.cell(row=current_row, column=7, value=row["company"])
        sheet.cell(row=current_row, column=8, value=row["phone"])

# Populate both sheets
populate_sheet(sheet_13, june13_rows, len(june13_rows))
populate_sheet(sheet_14, june14_rows, len(june14_rows))

# Save workbook
output_path = "/Users/nicksng/code/puzzle-portal/ocr_tools/results/attendance_split.xlsx"
wb.save(output_path)
print(f"Successfully saved split Excel file at: {output_path}")
