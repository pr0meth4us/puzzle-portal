import openpyxl
import re

# 1. Load Excel file
wb = openpyxl.load_workbook("/Users/nicksng/code/puzzle-portal/ocr_tools/results/sheet_export.xlsx")
sheet = wb["Actual Attendance MDK"]

# 2. Parse Markdown file
md_path = "/Users/nicksng/code/puzzle-portal/ocr_tools/results/attendance_transcription.md"
with open(md_path, "r", encoding="utf-8") as f:
    content = f.read()

rows_to_append = []
current_num = 93  # Start numbering from 93 (next row after 92 in the sheet)

for line in content.split("\n"):
    line = line.strip()
    if not line.startswith("|"):
        continue
    parts = [p.strip() for p in line.split("|")[1:-1]]
    if not parts or parts[0] == "ល.រ" or parts[0].startswith("---") or parts[0].startswith(":---"):
        continue
    
    name = parts[1]
    gender = parts[2]
    role = parts[3]
    company = parts[4]
    phone = parts[5]
    
    # Translate Gender to Status (ស្រី or ប្រុស)
    if gender in ["F", "F ", "ស្រី"]:
        status = "ស្រី"
    elif gender in ["M", "M ", "ប្រុស"]:
        status = "ប្រុស"
    else:
        status = gender
        
    row_data = {
        "num": current_num,
        "name": name,
        "status": status,
        "role": role if role not in ["None", ""] else None,
        "company": company if company not in ["None", ""] else None,
        "phone": phone if phone not in ["None", ""] else None
    }
    rows_to_append.append(row_data)
    current_num += 1

print(f"Parsed {len(rows_to_append)} rows from Markdown.")

# 3. Write rows to Excel sheet starting at Row 99
# Row 98 corresponds to guest No 92 (which is in row 98 of Excel)
# So guest No 93 goes to row 99 of Excel
start_row = 99
for i, row in enumerate(rows_to_append):
    current_row = start_row + i
    sheet.cell(row=current_row, column=2, value=row["num"])
    sheet.cell(row=current_row, column=3, value=row["name"])
    sheet.cell(row=current_row, column=4, value=None)  # Qty is blank
    sheet.cell(row=current_row, column=5, value=row["status"])
    sheet.cell(row=current_row, column=6, value=row["role"])
    sheet.cell(row=current_row, column=7, value=row["company"])
    sheet.cell(row=current_row, column=8, value=row["phone"])

# 4. Save workbook
output_path = "/Users/nicksng/code/puzzle-portal/ocr_tools/results/sheet_updated.xlsx"
wb.save(output_path)
print(f"Successfully created updated Excel sheet at: {output_path}")
